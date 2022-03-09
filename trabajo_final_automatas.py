import openpyxl
from openpyxl import Workbook
import re
import datetime
import time
from datetime import datetime
from time import mktime
# para escribir el excel
import xlsxwriter

while True:

    fecha_inicial=input("Digite fecha inicial en formato YYYY/MM/DD HH:MM : ")
    fecha_final=input("Digite fecha final en formato YYYY/MM/DD HH:MM : ")

    validate=re.search('[0-9]{4}/(0[1-9]|1[0-2])/(0[1-9]|[1-2][0-9]|3[0-1]) (2[0-3]|[01][0-9]):[0-5][0-9]',fecha_inicial)
    validate2=re.search('[0-9]{4}/(0[1-9]|1[0-2])/(0[1-9]|[1-2][0-9]|3[0-1]) (2[0-3]|[01][0-9]):[0-5][0-9]',fecha_final)
    if validate==None or validate2==None:
        print("\nError, formato de fecha/s incorrecto\n")

    else:
        # transformar str a date
        formatted_date1 = time.strptime(fecha_inicial, "%Y/%m/%d %H:%M")
        formatted_date2 = time.strptime(fecha_final, "%Y/%m/%d %H:%M")
        if formatted_date1 <= formatted_date2:
            print("\nFechas correctas. Procesando datos...\n")
            # pasamos de timestruct a datetime para coincidir con el formato xlsx
            fecha1 = datetime.fromtimestamp(mktime(formatted_date1))
            fecha2 = datetime.fromtimestamp(mktime(formatted_date2))
            break
        else:
            print("\nERROR:Fecha final menor a inicial\n")

# ABRIMOS EXCEL PARA IMPORTAR DATOS
# C45253
book = openpyxl.load_workbook('acts-user.xlsx')
ws = book['acts-user']

conexiones_iniciales = ws['C2':'C45253']
conexiones_finales=ws['D2':'D45253']

id_conexion=ws['A2':'A45253']
usuario=ws['B2':'B45253']
session_time=ws['E2':'E45253']
inputs=ws['F2':'F45253']
outputs=ws['G2':'G45253']
mac_ap=ws['H2':'H45253']
mac_client=ws['I2':'I45253']
# LLENAMOS LAS COLUMNAS Y LAS PONEMOS C/U EN UN ARRAY

array_inicio = []
array_fin=[]
array_id_conexion=[]
array_usuario=[]
array_session_time=[]
array_inputs=[]
array_outputs=[]
array_mac_ap=[]
array_mac_client=[]

# conexiones iniciales=45252
for fil in range(45252):
     for col in range(1):
        array_session_time.append(session_time[fil][col].value)
        array_inicio.append(conexiones_iniciales[fil][col].value)
        array_fin.append(conexiones_finales[fil][col].value)
        array_mac_client.append(mac_client[fil][col].value)
        array_usuario.append(usuario[fil][col].value)
        array_id_conexion.append(id_conexion[fil][col].value)
        array_inputs.append(inputs[fil][col].value)
        array_outputs.append(outputs[fil][col].value)
        array_mac_ap.append(mac_ap[fil][col].value)
        

contador=0
diccionario={}
array_diccionario=[]

for i in array_inicio:
    if i != None:
        if fecha1<=i and i<=fecha2:
            diccionario["Inicio"]=str(i)
            diccionario["Duracion"]=array_session_time[contador]
            diccionario["Fin de sesion"]=str(array_fin[contador])
            diccionario["Conexion ID"]=array_id_conexion[contador]
            diccionario["Usuario"]=array_usuario[contador]
            diccionario["Inputs"]=array_inputs[contador]
            diccionario["Outputs"]=array_outputs[contador]
            diccionario["Mac AP"]=array_mac_ap[contador]
            diccionario["Mac client"]=array_mac_client[contador]
            array_diccionario.append(diccionario)
            diccionario={}
    else:
        pass
    contador=contador+1
dic_ordenado=sorted(array_diccionario, key=lambda x: x['Duracion'],reverse=True)
for i in dic_ordenado:
    print(str(i)+"\n")

# Nuevo excel para insertar array de conexiones
workbook = xlsxwriter.Workbook("conexiones.xlsx")
worksheet = workbook.add_worksheet()

row = 0
headers=["Inicio","Duracion","Fin de sesion","Conexion ID","Usuario","Inputs","Outputs","Mac AP","Mac client"]

# recorre array dic_ordenado
for row, _dict in enumerate(dic_ordenado):
    # recorre llaves
    for col, key in enumerate(headers):
        worksheet.write(row, col, _dict[key])
workbook.close()